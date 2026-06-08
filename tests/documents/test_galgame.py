from __future__ import annotations

import base64
import json
from io import BytesIO
from pathlib import Path
from unittest.mock import MagicMock

import pytest
from openpyxl import Workbook, load_workbook

from context_aware_translation.documents.galgame import (
    GalgameDocument,
    KagKsScriptAdapter,
    MToolJsonAdapter,
    ParaTranzJsonAdapter,
    RenPyScriptAdapter,
    RpgMakerMvMzJsonAdapter,
    TppXlsxAdapter,
    TranslatorPlusPlusTransAdapter,
    VnTextJsonAdapter,
    WolfRpgXlsxAdapter,
    get_galgame_bridge_formats,
    get_galgame_external_tool_specs,
    inspect_galgame_external_helpers,
    inspect_galgame_glossary_seeds,
    inspect_galgame_import,
    validate_galgame_translations,
)
from context_aware_translation.utils.compression_marker import COMPRESSED_LINE_SENTINEL
from context_aware_translation.storage.repositories.document_repository import DocumentRepository
from context_aware_translation.storage.schema.book_db import SQLiteBookDB


def _setup_repo(tmp_path: Path) -> DocumentRepository:
    return DocumentRepository(SQLiteBookDB(tmp_path / "book.db"))


def _write_mtool_json(path: Path, mapping: dict[str, str]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(mapping, ensure_ascii=False), encoding="utf-8")


def _write_json(path: Path, payload: object) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(payload, ensure_ascii=False), encoding="utf-8")


def _workbook_source_text(workbook: Workbook) -> str:
    output = BytesIO()
    workbook.save(output)
    return base64.b64encode(output.getvalue()).decode("ascii")


def _write_workbook(path: Path, workbook: Workbook) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(path)


def _load_workbook_from_source_text(source_text: str) -> Workbook:
    return load_workbook(filename=BytesIO(base64.b64decode(source_text.encode("ascii"))))


def test_mtool_json_adapter_extracts_and_applies_translations() -> None:
    adapter = MToolJsonAdapter()
    source_text = json.dumps({"こんにちは": "", "またね": "old"}, ensure_ascii=False)

    units = adapter.extract_units("script.json", source_text)
    patched = json.loads(adapter.apply_translations("script.json", source_text, {"0": "你好"}))

    assert [(unit.unit_id, unit.text) for unit in units] == [("0", "こんにちは"), ("1", "またね")]
    assert patched == {"こんにちは": "你好", "またね": "old"}


def test_mtool_json_adapter_rejects_generic_config_json(tmp_path: Path) -> None:
    path = tmp_path / "config.json"
    _write_json(path, {"name": "game", "version": "1.0"})

    assert MToolJsonAdapter().can_read_file(path) is False
    assert GalgameDocument.can_import(path) is False


def test_galgame_phase6_bridge_and_external_tool_catalogs() -> None:
    bridges = get_galgame_bridge_formats()
    specs = get_galgame_external_tool_specs()

    assert any("translator_plus_plus_trans" in bridge.adapter_names for bridge in bridges)
    assert any(spec.name == "rpatool" and ".rpa" in spec.supported_extensions for spec in specs)


def test_inspect_galgame_external_helpers_plans_only_available_tools(tmp_path: Path) -> None:
    game = tmp_path / "game"
    archive = game / "archive.rpa"
    xp3_archive = game / "patch.xp3"
    unity_asset = game / "resources.assets"
    archive.parent.mkdir(parents=True)
    archive.write_bytes(b"rpa")
    xp3_archive.write_bytes(b"xp3")
    unity_asset.write_bytes(b"unity")

    summary = inspect_galgame_external_helpers(
        game,
        executable_resolver=lambda name: f"/tools/{name}" if name == "rpatool" else None,
    )

    assert summary.usable is True
    assert len(summary.commands) == 1
    command = summary.commands[0]
    assert command.tool_name == "rpatool"
    assert command.source_path == "archive.rpa"
    assert command.output_folder == "extracted/archive"
    assert command.argv == ("/tools/rpatool", "-x", archive.as_posix(), "-o", (game / "extracted" / "archive").as_posix())
    assert [(issue.source_path, issue.source_extension) for issue in summary.issues] == [
        ("patch.xp3", ".xp3"),
        ("resources.assets", ".assets"),
    ]


def test_translator_plus_plus_adapter_extracts_and_applies_translations(tmp_path: Path) -> None:
    adapter = TranslatorPlusPlusTransAdapter()
    source = {
        "project": {
            "files": {
                "script.ks": {
                    "data": [
                        ["こんにちは", ""],
                        {"original": "またね", "translation": "old", "rowInfoText": "scene 1"},
                        {"source": "選択肢", "target": ""},
                    ]
                }
            }
        }
    }
    source_text = json.dumps(source, ensure_ascii=False)
    path = tmp_path / "project.trans"
    path.write_text(source_text, encoding="utf-8")

    units = adapter.extract_units("project.trans", source_text)
    patched = json.loads(adapter.apply_translations("project.trans", source_text, {"0": "你好", "1": "再见", "2": "选项"}))

    assert adapter.can_read_file(path) is True
    assert [(unit.unit_id, unit.text, unit.context) for unit in units] == [
        ("0", "こんにちは", None),
        ("1", "またね", "scene 1"),
        ("2", "選択肢", None),
    ]
    rows = patched["project"]["files"]["script.ks"]["data"]
    assert rows[0][1] == "你好"
    assert rows[1]["translation"] == "再见"
    assert rows[2]["target"] == "选项"


def test_translator_plus_plus_can_read_file_rejects_empty_project(tmp_path: Path) -> None:
    path = tmp_path / "empty.trans"
    path.write_text(json.dumps({"project": {"files": {}}}), encoding="utf-8")

    assert TranslatorPlusPlusTransAdapter().can_read_file(path) is False


def test_vntext_adapter_extracts_and_applies_translations() -> None:
    adapter = VnTextJsonAdapter()
    source = [
        {"name": "Alice", "message": "こんにちは"},
        {"names": ["Bob", "Carol"], "message": "またね"},
        {"message": " "},
    ]
    source_text = json.dumps(source, ensure_ascii=False)

    units = adapter.extract_units("vntext.json", source_text)
    patched = json.loads(adapter.apply_translations("vntext.json", source_text, {"0": "你好", "1": "再见"}))

    assert [(unit.unit_id, unit.text, unit.speaker) for unit in units] == [
        ("0", "こんにちは", "Alice"),
        ("1", "またね", "Bob / Carol"),
    ]
    assert patched[0]["message"] == "你好"
    assert patched[1]["message"] == "再见"
    assert patched[2]["message"] == " "


def test_paratranz_adapter_extracts_and_applies_translations() -> None:
    adapter = ParaTranzJsonAdapter()
    source = [
        {"key": "line1", "original": "こんにちは", "translation": "", "speaker": "Alice", "context": "intro"},
        {"key": "line2", "original": "またね", "trans": "old", "name": "Bob"},
        {"key": "empty", "original": ""},
    ]
    source_text = json.dumps(source, ensure_ascii=False)

    units = adapter.extract_units("paratranz.json", source_text)
    patched = json.loads(adapter.apply_translations("paratranz.json", source_text, {"0": "你好", "1": "再见"}))

    assert [(unit.unit_id, unit.text, unit.speaker, unit.context) for unit in units] == [
        ("0", "こんにちは", "Alice", "intro"),
        ("1", "またね", "Bob", None),
    ]
    assert patched[0]["translation"] == "你好"
    assert patched[1]["trans"] == "再见"
    assert patched[2] == {"key": "empty", "original": ""}


def test_tpp_xlsx_adapter_extracts_and_applies_translations(tmp_path: Path) -> None:
    adapter = TppXlsxAdapter()
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Script"
    sheet.append(["Original Text", "Initial", "Name", "Context"])
    sheet.append(["こんにちは", "", "Alice", "intro"])
    sheet.append(["またね", "old", "Bob", "outro"])
    sheet.append(["", "ignored", "", ""])
    source_text = _workbook_source_text(workbook)
    path = tmp_path / "script.xlsx"
    _write_workbook(path, workbook)

    units = adapter.extract_units("script.xlsx", source_text)
    patched_workbook = _load_workbook_from_source_text(
        adapter.apply_translations("script.xlsx", source_text, {"0": "你好", "1": "再见"})
    )

    assert adapter.can_read_file(path) is True
    assert [(unit.unit_id, unit.text, unit.speaker, unit.context) for unit in units] == [
        ("0", "こんにちは", "Alice", "intro"),
        ("1", "またね", "Bob", "outro"),
    ]
    patched_sheet = patched_workbook["Script"]
    assert patched_sheet.cell(row=2, column=2).value == "你好"
    assert patched_sheet.cell(row=3, column=2).value == "再见"


def test_tpp_xlsx_adapter_uses_deterministic_header_priority() -> None:
    adapter = TppXlsxAdapter()
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["Source", "Original Text", "Target", "Initial"])
    sheet.append(["wrong source", "こんにちは", "wrong target", ""])
    source_text = _workbook_source_text(workbook)

    units = adapter.extract_units("script.xlsx", source_text)
    patched_workbook = _load_workbook_from_source_text(adapter.apply_translations("script.xlsx", source_text, {"0": "你好"}))

    assert [unit.text for unit in units] == ["こんにちは"]
    patched_sheet = patched_workbook.active
    assert patched_sheet.cell(row=2, column=3).value == "wrong target"
    assert patched_sheet.cell(row=2, column=4).value == "你好"


def test_wolf_rpg_xlsx_adapter_extracts_and_applies_translations() -> None:
    adapter = WolfRpgXlsxAdapter()
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Map001"
    sheet.append(["文件", "原文", "译文", "角色", "备注"])
    sheet.append(["Map001.json", "こんにちは", "", "Alice", "event 1"])
    sheet.append(["Map001.json", "またね", "old", "Bob", "event 2"])
    source_text = _workbook_source_text(workbook)

    units = adapter.extract_units("wolf.xlsx", source_text)
    patched_workbook = _load_workbook_from_source_text(
        adapter.apply_translations("wolf.xlsx", source_text, {"0": "你好", "1": "再见"})
    )

    assert [(unit.unit_id, unit.text, unit.speaker, unit.context) for unit in units] == [
        ("0", "こんにちは", "Alice", "event 1"),
        ("1", "またね", "Bob", "event 2"),
    ]
    patched_sheet = patched_workbook["Map001"]
    assert patched_sheet.cell(row=2, column=3).value == "你好"
    assert patched_sheet.cell(row=3, column=3).value == "再见"


def test_renpy_adapter_extracts_and_patches_native_script() -> None:
    adapter = RenPyScriptAdapter()
    source_text = """label start:\n    e \"こんにちは\"\n    \"またね\"\n    menu:\n        \"行く\":\n            jump go\n    voice \"line.ogg\"\n"""

    units = adapter.extract_units("script.rpy", source_text)
    patched = adapter.apply_translations("script.rpy", source_text, {"0": "你好", "1": "再见", "2": "走"})

    assert [(unit.unit_id, unit.text, unit.speaker) for unit in units] == [
        ("0", "こんにちは", "e"),
        ("1", "またね", None),
        ("2", "行く", None),
    ]
    assert "    e \"你好\"\n" in patched
    assert "    \"再见\"\n" in patched
    assert "        \"走\":\n" in patched
    assert "    voice \"line.ogg\"\n" in patched


def test_renpy_adapter_patches_generated_old_new_translation_blocks() -> None:
    adapter = RenPyScriptAdapter()
    source_text = """translate japanese strings:\n    old \"Start Game\"\n    new \"\"\n"""

    units = adapter.extract_units("translations.rpy", source_text)
    patched = adapter.apply_translations("translations.rpy", source_text, {"0": "开始游戏"})

    assert [(unit.text, unit.metadata["line"], unit.metadata["target_line"]) for unit in units] == [
        ("Start Game", 2, 3)
    ]
    assert "    old \"Start Game\"\n" in patched
    assert "    new \"开始游戏\"\n" in patched


def test_kag_ks_adapter_extracts_and_patches_dialogue_lines() -> None:
    adapter = KagKsScriptAdapter()
    source_text = """*start\n#Alice\nこんにちは[l][r]\n@bg storage=room\n[link target=*yes]はい[endlink]\n; comment\n"""

    units = adapter.extract_units("scene.ks", source_text)
    patched = adapter.apply_translations("scene.ks", source_text, {"0": "你好[l][r]", "1": "[link target=*yes]是[endlink]"})

    assert [(unit.unit_id, unit.text, unit.speaker) for unit in units] == [
        ("0", "こんにちは[l][r]", "Alice"),
        ("1", "[link target=*yes]はい[endlink]", "Alice"),
    ]
    assert "こんにちは" not in patched
    assert "你好[l][r]\n" in patched
    assert "@bg storage=room\n" in patched
    assert "[link target=*yes]是[endlink]\n" in patched


def test_rpg_maker_mv_mz_adapter_extracts_and_patches_map_events() -> None:
    adapter = RpgMakerMvMzJsonAdapter()
    source = {
        "events": [
            None,
            {
                "name": "EV001",
                "pages": [
                    {
                        "list": [
                            {"code": 101, "parameters": ["Actor1", 0, 0, 2, "アリス"]},
                            {"code": 401, "parameters": ["こんにちは\\N[1]"]},
                            {"code": 102, "parameters": [["はい", "いいえ"], 0, 0, 2, 0]},
                            {"code": 356, "parameters": ["PluginCommand should stay"]},
                        ]
                    }
                ],
            },
        ]
    }
    source_text = json.dumps(source, ensure_ascii=False)

    units = adapter.extract_units("data/Map001.json", source_text)
    patched = json.loads(
        adapter.apply_translations(
            "data/Map001.json",
            source_text,
            {"0": "Alice", "1": "你好\\N[1]", "2": "是", "3": "不是"},
        )
    )

    assert [(unit.unit_id, unit.text, unit.speaker, unit.metadata["kind"]) for unit in units] == [
        ("0", "アリス", "アリス", "speaker"),
        ("1", "こんにちは\\N[1]", "アリス", "dialogue"),
        ("2", "はい", None, "choice"),
        ("3", "いいえ", None, "choice"),
    ]
    commands = patched["events"][1]["pages"][0]["list"]
    assert commands[0]["parameters"][4] == "Alice"
    assert commands[1]["parameters"][0] == "你好\\N[1]"
    assert commands[2]["parameters"][0] == ["是", "不是"]
    assert commands[3]["parameters"][0] == "PluginCommand should stay"


def test_rpg_maker_adapter_rejects_unrelated_json_path() -> None:
    adapter = RpgMakerMvMzJsonAdapter()
    source_text = json.dumps({"events": []}, ensure_ascii=False)

    with pytest.raises(ValueError, match="Unsupported RPG Maker"):
        adapter.extract_units("config/settings.json", source_text)


def test_inspect_galgame_import_reports_summary_and_manual_adapter_override(tmp_path: Path) -> None:
    source = tmp_path / "game" / "data" / "Map001.json"
    _write_json(source, {"events": [None, {"pages": [{"list": [{"code": 401, "parameters": ["こんにちは"]}]}]}]})

    summary = inspect_galgame_import(source.parent.parent)
    overridden_summary = inspect_galgame_import(source.parent.parent, adapter_name="mtool_json")

    assert summary.imported == 1
    assert summary.total_units == 1
    assert summary.adapters == ("rpg_maker_mv_mz_json",)
    assert summary.files[0].relative_path == "data/Map001.json"
    assert summary.files[0].confidence == "high"
    assert overridden_summary.imported == 0


def test_inspect_galgame_import_reports_medium_confidence_for_single_rpg_maker_file(tmp_path: Path) -> None:
    source = tmp_path / "Map001.json"
    _write_json(source, {"events": [None, {"pages": [{"list": [{"code": 401, "parameters": ["こんにちは"]}]}]}]})

    summary = inspect_galgame_import(source)

    assert summary.imported == 1
    assert summary.files[0].relative_path == "Map001.json"
    assert summary.files[0].adapter_name == "rpg_maker_mv_mz_json"
    assert summary.files[0].confidence == "medium"


def test_inspect_galgame_glossary_seeds_reads_units_and_project_metadata(tmp_path: Path) -> None:
    script = tmp_path / "game" / "script.rpy"
    script.parent.mkdir(parents=True)
    script.write_text('define e = Character("エイリーン")\nlabel start:\n    e "こんにちは"\n', encoding="utf-8")
    _write_json(
        tmp_path / "game" / "data" / "Actors.json",
        [None, {"id": 1, "name": "アリス", "nickname": "勇者"}],
    )
    _write_json(tmp_path / "game" / "data" / "System.json", {"gameTitle": "月の物語", "currencyUnit": "G"})
    _write_json(
        tmp_path / "game" / "data" / "Map001.json",
        {"events": [None, {"pages": [{"list": [{"code": 101, "parameters": ["", 0, 0, 2, "ボブ"]}]}]}]},
    )

    seeds = inspect_galgame_glossary_seeds(tmp_path / "game")

    assert {(seed.term, seed.category, seed.source) for seed in seeds} == {
        ("エイリーン", "character", "script.rpy"),
        ("e", "speaker", "script.rpy"),
        ("アリス", "character", "data/Actors.json"),
        ("勇者", "character", "data/Actors.json"),
        ("月の物語", "project", "data/System.json"),
        ("G", "term", "data/System.json"),
        ("ボブ", "speaker", "data/Map001.json"),
    }


def test_galgame_can_import_mtool_json(tmp_path: Path) -> None:
    source = tmp_path / "script.json"
    _write_mtool_json(source, {"こんにちは": ""})

    assert GalgameDocument.can_import(source) is True

    source.write_text("{}", encoding="utf-8")
    assert GalgameDocument.can_import(source) is False


async def test_galgame_document_preserve_export_uses_imported_json_adapter_mime_type(tmp_path: Path) -> None:
    source = tmp_path / "vntext.json"
    _write_json(source, [{"name": "Alice", "message": "こんにちは"}])
    repo = _setup_repo(tmp_path)

    result = GalgameDocument.do_import(repo, source)

    assert result == {"imported": 1, "skipped": 0}
    row = repo.get_document_row()
    assert row is not None
    document = GalgameDocument(repo, row["document_id"])
    sources = repo.get_document_sources(row["document_id"])
    assert sources[0]["mime_type"] == VnTextJsonAdapter.mime_type

    assert document.get_text().splitlines() == ["こんにちは"]

    await document.set_text(["你好"])
    output_folder = tmp_path / "out"
    document.export_preserve_structure(output_folder)

    patched = json.loads((output_folder / "vntext.json").read_text(encoding="utf-8"))
    assert patched == [{"name": "Alice", "message": "你好"}]


async def test_galgame_document_preserve_export_writes_imported_xlsx_adapter(tmp_path: Path) -> None:
    source = tmp_path / "script.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Script"
    sheet.append(["Original Text", "Initial", "Name"])
    sheet.append(["こんにちは", "", "Alice"])
    _write_workbook(source, workbook)
    repo = _setup_repo(tmp_path)

    result = GalgameDocument.do_import(repo, source)

    assert result == {"imported": 1, "skipped": 0}
    row = repo.get_document_row()
    assert row is not None
    document = GalgameDocument(repo, row["document_id"])
    sources = repo.get_document_sources(row["document_id"])
    assert sources[0]["mime_type"] == TppXlsxAdapter.mime_type

    assert document.get_text().splitlines() == ["こんにちは"]

    await document.set_text(["你好"])
    output_folder = tmp_path / "out"
    document.export_preserve_structure(output_folder)

    patched = load_workbook(output_folder / "script.xlsx")
    assert patched["Script"].cell(row=2, column=2).value == "你好"


async def test_galgame_document_imports_and_exports_renpy_script(tmp_path: Path) -> None:
    source = tmp_path / "game" / "script.rpy"
    source.parent.mkdir(parents=True, exist_ok=True)
    source.write_text('label start:\n    e "こんにちは"\n', encoding="utf-8")
    repo = _setup_repo(tmp_path)

    result = GalgameDocument.do_import(repo, source.parent)

    assert result == {"imported": 1, "skipped": 0}
    row = repo.get_document_row()
    assert row is not None
    document = GalgameDocument(repo, row["document_id"])
    sources = repo.get_document_sources(row["document_id"])
    assert sources[0]["mime_type"] == RenPyScriptAdapter.mime_type

    assert document.get_text().splitlines() == ["こんにちは"]

    await document.set_text(["你好"])
    output_folder = tmp_path / "out"
    document.export_preserve_structure(output_folder)

    assert (output_folder / "script.rpy").read_text(encoding="utf-8") == 'label start:\n    e "你好"\n'


def test_galgame_document_reports_glossary_seeds_from_imported_sources(tmp_path: Path) -> None:
    source = tmp_path / "game" / "script.rpy"
    source.parent.mkdir(parents=True, exist_ok=True)
    source.write_text('define e = Character("エイリーン")\nlabel start:\n    e "こんにちは"\n', encoding="utf-8")
    repo = _setup_repo(tmp_path)

    GalgameDocument.do_import(repo, source.parent)
    row = repo.get_document_row()
    assert row is not None
    document = GalgameDocument(repo, row["document_id"])

    seeds = document.get_glossary_seeds()

    assert [(seed.term, seed.category, seed.source) for seed in seeds] == [
        ("e", "speaker", "script.rpy"),
        ("エイリーン", "character", "script.rpy"),
    ]


async def test_galgame_document_imports_and_exports_rpg_maker_folder(tmp_path: Path) -> None:
    source = tmp_path / "game" / "www" / "data" / "Map001.json"
    _write_json(source, {"events": [None, {"pages": [{"list": [{"code": 401, "parameters": ["こんにちは"]}]}]}]})
    repo = _setup_repo(tmp_path)

    result = GalgameDocument.do_import(repo, source.parents[2])

    assert result == {"imported": 1, "skipped": 0}
    row = repo.get_document_row()
    assert row is not None
    document = GalgameDocument(repo, row["document_id"])
    sources = repo.get_document_sources(row["document_id"])
    assert sources[0]["mime_type"] == RpgMakerMvMzJsonAdapter.mime_type

    assert document.get_text().splitlines() == ["こんにちは"]

    await document.set_text(["你好"])
    output_folder = tmp_path / "out"
    document.export_preserve_structure(output_folder)

    patched = json.loads((output_folder / "www" / "data" / "Map001.json").read_text(encoding="utf-8"))
    assert patched["events"][1]["pages"][0]["list"][0]["parameters"] == ["你好"]


async def test_galgame_document_keeps_compressed_line_as_real_line(tmp_path: Path) -> None:
    source = tmp_path / "script.json"
    _write_mtool_json(source, {"こんにちは": ""})
    repo = _setup_repo(tmp_path)

    GalgameDocument.do_import(repo, source)
    row = repo.get_document_row()
    assert row is not None
    document = GalgameDocument(repo, row["document_id"])

    await document.set_text([COMPRESSED_LINE_SENTINEL])

    assert document._translated_lines == [COMPRESSED_LINE_SENTINEL]


async def test_galgame_document_imports_single_rpg_maker_map_file(tmp_path: Path) -> None:
    source = tmp_path / "Map001.json"
    _write_json(source, {"events": [None, {"pages": [{"list": [{"code": 401, "parameters": ["こんにちは"]}]}]}]})
    repo = _setup_repo(tmp_path)

    assert RpgMakerMvMzJsonAdapter().can_read_file(source) is True
    result = GalgameDocument.do_import(repo, source)

    assert result == {"imported": 1, "skipped": 0}
    row = repo.get_document_row()
    assert row is not None
    document = GalgameDocument(repo, row["document_id"])
    sources = repo.get_document_sources(row["document_id"])
    assert sources[0]["relative_path"] == "Map001.json"
    assert sources[0]["mime_type"] == RpgMakerMvMzJsonAdapter.mime_type
    assert document.get_text().splitlines() == ["こんにちは"]


def test_galgame_folder_scan_rejects_rpg_maker_named_json_outside_data_folder(tmp_path: Path) -> None:
    source = tmp_path / "config" / "Map001.json"
    _write_json(source, {"events": [None, {"pages": [{"list": [{"code": 401, "parameters": ["こんにちは"]}]}]}]})

    assert GalgameDocument.can_import(tmp_path) is False


async def test_galgame_merged_export_is_rejected(tmp_path: Path) -> None:
    source = tmp_path / "script.json"
    _write_mtool_json(source, {"こんにちは": ""})
    repo = _setup_repo(tmp_path)
    GalgameDocument.do_import(repo, source)
    row = repo.get_document_row()
    assert row is not None
    document = GalgameDocument(repo, row["document_id"])
    await document.set_text(["你好"])

    assert document.can_export("native") is True
    assert document.can_export("json") is False
    assert document.can_export("txt") is False
    with pytest.raises(ValueError, match="preserve-structure export"):
        GalgameDocument.export_merged([document], "native", tmp_path / "out")


async def test_galgame_preserve_export_writes_single_imported_json_source(tmp_path: Path) -> None:
    source = tmp_path / "script.json"
    _write_mtool_json(source, {"こんにちは": ""})
    repo = _setup_repo(tmp_path)
    GalgameDocument.do_import(repo, source)
    row = repo.get_document_row()
    assert row is not None
    document = GalgameDocument(repo, row["document_id"])
    await document.set_text(["你好"])

    output_folder = tmp_path / "out"
    document.export_preserve_structure(output_folder)

    assert json.loads((output_folder / "script.json").read_text(encoding="utf-8")) == {"こんにちは": "你好"}


async def test_galgame_preserve_export_preserves_imported_json_adapter_shape(tmp_path: Path) -> None:
    cases = [
        (
            "vntext.json",
            [{"name": "Alice", "message": "こんにちは"}],
            VnTextJsonAdapter.mime_type,
            lambda patched: patched == [{"name": "Alice", "message": "你好"}],
        ),
        (
            "paratranz.json",
            [{"key": "line1", "original": "こんにちは", "translation": ""}],
            ParaTranzJsonAdapter.mime_type,
            lambda patched: patched == [{"key": "line1", "original": "こんにちは", "translation": "你好"}],
        ),
        (
            "Map001.json",
            {"events": [None, {"pages": [{"list": [{"code": 401, "parameters": ["こんにちは"]}]}]}]},
            RpgMakerMvMzJsonAdapter.mime_type,
            lambda patched: patched["events"][1]["pages"][0]["list"][0]["parameters"] == ["你好"],
        ),
    ]

    for filename, payload, expected_mime_type, assert_patched in cases:
        case_dir = tmp_path / Path(filename).stem
        source = case_dir / filename
        _write_json(source, payload)
        repo = _setup_repo(case_dir)
        GalgameDocument.do_import(repo, source)
        row = repo.get_document_row()
        assert row is not None
        sources = repo.get_document_sources(row["document_id"])
        assert sources[0]["mime_type"] == expected_mime_type
        document = GalgameDocument(repo, row["document_id"])
        await document.set_text(["你好"])

        output_folder = case_dir / "out"
        document.export_preserve_structure(output_folder)

        assert assert_patched(json.loads((output_folder / filename).read_text(encoding="utf-8")))


async def test_galgame_preserve_export_writes_imported_trans_source(tmp_path: Path) -> None:
    source = tmp_path / "project.trans"
    _write_json(
        source,
        {
            "project": {
                "files": {
                    "script.ks": {
                        "data": [
                            ["こんにちは", ""],
                            {"original": "またね", "translation": "old"},
                        ]
                    }
                }
            }
        },
    )
    repo = _setup_repo(tmp_path)
    GalgameDocument.do_import(repo, source)
    row = repo.get_document_row()
    assert row is not None
    document = GalgameDocument(repo, row["document_id"])
    await document.set_text(["你好", "再见"])

    output_folder = tmp_path / "out"
    document.export_preserve_structure(output_folder)

    rows = json.loads((output_folder / "project.trans").read_text(encoding="utf-8"))["project"]["files"]["script.ks"]["data"]
    assert rows[0][1] == "你好"
    assert rows[1]["translation"] == "再见"


async def test_galgame_preserve_export_writes_imported_renpy_source(tmp_path: Path) -> None:
    source = tmp_path / "script.rpy"
    source.write_text('label start:\n    e "こんにちは"\n', encoding="utf-8")
    repo = _setup_repo(tmp_path)
    GalgameDocument.do_import(repo, source)
    row = repo.get_document_row()
    assert row is not None
    document = GalgameDocument(repo, row["document_id"])
    await document.set_text(["你好"])

    output_folder = tmp_path / "out"
    document.export_preserve_structure(output_folder)

    assert (output_folder / "script.rpy").read_text(encoding="utf-8") == 'label start:\n    e "你好"\n'


async def test_galgame_document_import_get_text_and_preserve_export(tmp_path: Path) -> None:
    source = tmp_path / "game" / "data" / "script.json"
    _write_mtool_json(source, {"こんにちは": "", "またね": ""})
    repo = _setup_repo(tmp_path)

    result = GalgameDocument.do_import(repo, source.parent.parent)

    assert result == {"imported": 1, "skipped": 0}
    row = repo.get_document_row()
    assert row is not None
    assert row["document_type"] == "galgame"
    sources = repo.get_document_sources(row["document_id"])
    assert sources[0]["relative_path"] == "data/script.json"

    document = GalgameDocument(repo, row["document_id"])
    assert document.get_text().splitlines() == ["こんにちは", "またね"]

    await document.set_text(["你好", "再见"])
    output_folder = tmp_path / "out"
    document.export_preserve_structure(output_folder)

    patched = json.loads((output_folder / "data" / "script.json").read_text(encoding="utf-8"))
    assert patched == {"こんにちは": "你好", "またね": "再见"}


async def test_galgame_document_rejects_grouped_chunk_translation_lines(tmp_path: Path) -> None:
    source = tmp_path / "script.json"
    _write_mtool_json(source, {"こんにちは": "", "またね": ""})
    repo = _setup_repo(tmp_path)
    GalgameDocument.do_import(repo, source)
    row = repo.get_document_row()
    assert row is not None
    document = GalgameDocument(repo, row["document_id"])

    with pytest.raises(ValueError, match="line stream entries cannot contain newline"):
        await document.set_text(["你好\n再见"])


async def test_galgame_document_rejects_unit_aligned_embedded_newline_translation(tmp_path: Path) -> None:
    source = tmp_path / "script.json"
    _write_mtool_json(source, {"こんにちは": "", "またね": ""})
    repo = _setup_repo(tmp_path)
    GalgameDocument.do_import(repo, source)
    row = repo.get_document_row()
    assert row is not None
    document = GalgameDocument(repo, row["document_id"])

    with pytest.raises(ValueError, match="line stream entries cannot contain newline"):
        await document.set_text(["你好\n补一句", "再见"])


async def test_galgame_document_splits_multiline_units_by_original_line_counts(tmp_path: Path) -> None:
    source = tmp_path / "script.json"
    _write_mtool_json(source, {"こんにちは\n世界": "", "またね": ""})
    repo = _setup_repo(tmp_path)
    GalgameDocument.do_import(repo, source)
    row = repo.get_document_row()
    assert row is not None
    document = GalgameDocument(repo, row["document_id"])

    await document.set_text(["你好", "世界", "再见"])
    output_folder = tmp_path / "out"
    document.export_preserve_structure(output_folder)

    patched = json.loads((output_folder / "script.json").read_text(encoding="utf-8"))
    assert patched == {"こんにちは\n世界": "你好\n世界", "またね": "再见"}


async def test_galgame_document_rejects_unit_count_aligned_but_short_physical_line_stream(tmp_path: Path) -> None:
    source = tmp_path / "script.json"
    _write_mtool_json(source, {"こんにちは\n世界": "", "またね": ""})
    repo = _setup_repo(tmp_path)
    GalgameDocument.do_import(repo, source)
    row = repo.get_document_row()
    assert row is not None
    document = GalgameDocument(repo, row["document_id"])

    with pytest.raises(ValueError, match="shorter than the source unit stream"):
        await document.set_text(["你好", "再见"])


async def test_galgame_export_raises_on_short_translation_stream() -> None:
    mock_repo = MagicMock()
    mock_repo.get_document_sources.return_value = [
        {
            "sequence_number": 0,
            "relative_path": "script.json",
            "text_content": json.dumps({"こんにちは": "", "またね": ""}, ensure_ascii=False),
            "mime_type": MToolJsonAdapter.mime_type,
        }
    ]
    document = GalgameDocument(mock_repo, 1)

    with pytest.raises(ValueError, match="shorter than the source unit stream"):
        await document.set_text(["你好"])
