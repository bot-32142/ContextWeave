from __future__ import annotations

import base64
import binascii
import json
import re
import shutil
from collections.abc import Callable, Iterable, Mapping
from dataclasses import dataclass, field
from io import BytesIO
from pathlib import Path, PurePosixPath
from typing import TYPE_CHECKING, Any, Protocol, cast
from zipfile import BadZipFile

from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException

from context_aware_translation.core.cancellation import raise_if_cancelled
from context_aware_translation.core.progress import ProgressCallback
from context_aware_translation.documents.base import Document
from context_aware_translation.utils.compression_marker import decode_compressed_line

if TYPE_CHECKING:
    from context_aware_translation.config import ImageReembeddingConfig
    from context_aware_translation.llm.client import LLMClient
    from context_aware_translation.storage.repositories.document_repository import DocumentRepository


_SIMPLE_IDENTIFIER_RE = re.compile(r"^[A-Za-z_][A-Za-z0-9_.:-]{0,40}$")
_KAG_TAG_RE = re.compile(r"\[[^\]]*\]")
_RPG_MAKER_MAP_FILE_RE = re.compile(r"^Map\d+\.json$", re.IGNORECASE)
_LINE_CONSTRAINED_ADAPTERS = frozenset({"renpy_rpy", "kag_ks", "rpg_maker_mv_mz_json"})
_RENPY_CHARACTER_DEF_RE = re.compile(r"^\s*(?:define\s+)?(?P<symbol>[A-Za-z_]\w*)\s*=\s*(?:renpy\.)?Character\s*\(")
_EXTERNAL_ARCHIVE_EXTENSIONS = frozenset({".rpa", ".rpyc", ".xp3", ".assets", ".unity3d", ".bundle", ".rvdata2", ".rxdata"})
_RENPY_COMMAND_PREFIXES = frozenset(
    {
        "call",
        "define",
        "default",
        "elif",
        "else",
        "for",
        "hide",
        "if",
        "image",
        "init",
        "jump",
        "label",
        "menu",
        "pause",
        "play",
        "python",
        "return",
        "scene",
        "screen",
        "show",
        "stop",
        "transform",
        "translate",
        "voice",
        "while",
        "with",
        "window",
    }
)


@dataclass(frozen=True)
class TranslationUnit:
    """One translatable galgame script/intermediate-file unit."""

    relative_path: str
    unit_id: str
    text: str
    speaker: str | None = None
    context: str | None = None
    metadata: Mapping[str, object] = field(default_factory=dict)


@dataclass(frozen=True)
class _SpreadsheetColumns:
    header_row: int
    source_column: int
    translation_column: int | None
    speaker_column: int | None
    context_column: int | None


@dataclass(frozen=True)
class _QuotedString:
    start_quote: int
    end_quote: int
    quote: str
    text: str


@dataclass(frozen=True)
class _RenPyLine:
    kind: str
    text: str
    speaker: str | None


@dataclass(frozen=True)
class GalgameImportSourceSummary:
    """One importable galgame source file discovered during inspection."""

    relative_path: str
    adapter_name: str
    unit_count: int
    confidence: str


@dataclass(frozen=True)
class GalgameImportSummary:
    """Adapter-level import summary for UI and CLI preview flows."""

    files: tuple[GalgameImportSourceSummary, ...]
    skipped: int = 0

    @property
    def imported(self) -> int:
        return len(self.files)

    @property
    def total_units(self) -> int:
        return sum(source.unit_count for source in self.files)

    @property
    def adapters(self) -> tuple[str, ...]:
        return tuple(dict.fromkeys(source.adapter_name for source in self.files))


@dataclass(frozen=True)
class GalgameValidationIssue:
    """A safety issue found before patching translated galgame text."""

    relative_path: str
    unit_id: str
    message: str


@dataclass(frozen=True)
class GalgameBridgeFormat:
    """Offline bridge format produced by an external extractor or translation tool."""

    tool_name: str
    adapter_names: tuple[str, ...]
    output_extensions: tuple[str, ...]
    note: str


@dataclass(frozen=True)
class GalgameExternalToolSpec:
    """Known external helper that can unpack a format into editable files."""

    name: str
    executable_names: tuple[str, ...]
    supported_extensions: tuple[str, ...]
    purpose: str
    note: str


@dataclass(frozen=True)
class GalgameExternalHelperCommand:
    """A dry-run command for an explicit external extraction helper."""

    source_path: str
    output_folder: str
    tool_name: str
    executable: str
    argv: tuple[str, ...]
    note: str


@dataclass(frozen=True)
class GalgameExternalHelperIssue:
    """A file that needs external extraction but has no usable helper command."""

    source_path: str
    source_extension: str
    message: str


@dataclass(frozen=True)
class GalgameExternalHelperSummary:
    """Dry-run summary for external archive/script helper integration."""

    commands: tuple[GalgameExternalHelperCommand, ...]
    issues: tuple[GalgameExternalHelperIssue, ...]

    @property
    def usable(self) -> bool:
        return bool(self.commands)


@dataclass(frozen=True)
class GalgameGlossarySeed:
    """Glossary seed extracted from galgame speakers, character definitions, or metadata."""

    term: str
    category: str
    source: str
    context: str | None = None


@dataclass
class _RpgMakerTextSlot:
    text: str
    speaker: str | None
    context: str
    metadata: dict[str, object]
    parameters: list[object]
    parameter_index: int
    choice_index: int | None = None


class GalgameAdapter(Protocol):
    """Adapter for one galgame script or extracted/intermediate file format."""

    name: str
    mime_type: str
    supported_extensions: frozenset[str]
    stores_binary_as_base64: bool

    def can_read_file(self, path: Path) -> bool:
        """Return true when this adapter can extract units from a file."""
        ...

    def extract_units(self, relative_path: str, text: str) -> list[TranslationUnit]:
        """Extract translatable units from one source file."""
        ...

    def apply_translations(self, relative_path: str, text: str, translations: Mapping[str, str]) -> str:
        """Patch translated unit text back into the original source file text."""
        ...


_BRIDGE_FORMATS = (
    GalgameBridgeFormat(
        tool_name="MTool",
        adapter_names=("mtool_json",),
        output_extensions=(".json",),
        note="Use MTool to extract game text, then import the produced JSON dictionary.",
    ),
    GalgameBridgeFormat(
        tool_name="Translator++",
        adapter_names=("translator_plus_plus_trans", "tpp_xlsx"),
        output_extensions=(".trans", ".xlsx"),
        note="Import Translator++ project files or spreadsheet exports instead of patching archives directly.",
    ),
    GalgameBridgeFormat(
        tool_name="VNText",
        adapter_names=("vntext_json",),
        output_extensions=(".json",),
        note="Import VNText JSON exports containing message/name entries.",
    ),
    GalgameBridgeFormat(
        tool_name="ParaTranz",
        adapter_names=("paratranz_json",),
        output_extensions=(".json",),
        note="Import ParaTranz JSON files with original/translation entries.",
    ),
    GalgameBridgeFormat(
        tool_name="Wolf RPG extraction tools",
        adapter_names=("wolf_rpg_xlsx",),
        output_extensions=(".xlsx",),
        note="Import extracted Wolf RPG spreadsheets instead of binary project data.",
    ),
)

_EXTERNAL_TOOL_SPECS = (
    GalgameExternalToolSpec(
        name="rpatool",
        executable_names=("rpatool",),
        supported_extensions=(".rpa",),
        purpose="Extract Ren'Py `.rpa` archives into editable project files.",
        note="Run explicitly on a copy of the project, then import the extracted `.rpy` or JSON files.",
    ),
    GalgameExternalToolSpec(
        name="unrpyc",
        executable_names=("unrpyc",),
        supported_extensions=(".rpyc",),
        purpose="Decompile Ren'Py `.rpyc` bytecode into `.rpy` scripts.",
        note="Review decompiled scripts before import; generated code may need cleanup.",
    ),
    GalgameExternalToolSpec(
        name="garbro",
        executable_names=("GARbro.Console", "garbro"),
        supported_extensions=(".xp3",),
        purpose="Extract KiriKiri `.xp3` archives into editable files.",
        note="Run explicitly on a copy of the archive, then import extracted `.ks` files.",
    ),
)


class MToolJsonAdapter:
    """Adapter for MTool-style JSON dictionaries: original text -> translated text."""

    name = "mtool_json"
    mime_type = "application/x-contextweave-galgame-mtool-json"
    supported_extensions = frozenset({".json"})
    stores_binary_as_base64 = False

    def can_read_file(self, path: Path) -> bool:
        if path.suffix.lower() not in self.supported_extensions or not path.is_file():
            return False
        try:
            text = path.read_text(encoding="utf-8-sig")
        except UnicodeDecodeError:
            return False
        return self._load_mapping(text) is not None

    def extract_units(self, relative_path: str, text: str) -> list[TranslationUnit]:
        mapping = self._load_mapping(text)
        if mapping is None:
            raise ValueError(f"Unsupported MTool JSON file: {relative_path}")

        units: list[TranslationUnit] = []
        for index, source_text in enumerate(mapping):
            if not source_text.strip():
                continue
            units.append(
                TranslationUnit(
                    relative_path=relative_path,
                    unit_id=str(index),
                    text=source_text,
                    metadata={"adapter": self.name, "json_key": source_text},
                )
            )
        return units

    def apply_translations(self, relative_path: str, text: str, translations: Mapping[str, str]) -> str:
        mapping = self._load_mapping(text)
        if mapping is None:
            raise ValueError(f"Unsupported MTool JSON file: {relative_path}")

        patched: dict[str, str] = {}
        for index, (source_text, existing_translation) in enumerate(mapping.items()):
            patched[source_text] = _decode_compressed_line_stream(translations.get(str(index), existing_translation))
        return json.dumps(patched, ensure_ascii=False, indent=2) + "\n"

    @staticmethod
    def _load_mapping(text: str) -> dict[str, str] | None:
        try:
            loaded = json.loads(text)
        except json.JSONDecodeError:
            return None
        if not isinstance(loaded, dict) or not loaded:
            return None

        mapping: dict[str, str] = {}
        for key, value in loaded.items():
            if not isinstance(key, str) or not isinstance(value, str):
                return None
            if key.strip():
                mapping[key] = value
        if not mapping:
            return None
        if not any(_looks_like_translatable_json_key(source_text) for source_text in mapping):
            return None
        return mapping


class TranslatorPlusPlusTransAdapter:
    """Adapter for Translator++ `.trans` project files."""

    name = "translator_plus_plus_trans"
    mime_type = "application/x-contextweave-galgame-translator-plus-plus"
    supported_extensions = frozenset({".trans"})
    stores_binary_as_base64 = False

    def can_read_file(self, path: Path) -> bool:
        if path.suffix.lower() not in self.supported_extensions or not path.is_file():
            return False
        try:
            text = path.read_text(encoding="utf-8-sig")
        except UnicodeDecodeError:
            return False
        try:
            return bool(self.extract_units(path.name, text))
        except ValueError:
            return False

    def extract_units(self, relative_path: str, text: str) -> list[TranslationUnit]:
        root = self._load_root(text)
        if root is None:
            raise ValueError(f"Unsupported Translator++ project file: {relative_path}")

        units: list[TranslationUnit] = []
        for file_key, data_rows in self._iter_data_rows(root):
            for row_index, row in data_rows:
                source_text = self._row_source_text(row)
                if source_text is None or not source_text.strip():
                    continue
                units.append(
                    TranslationUnit(
                        relative_path=relative_path,
                        unit_id=str(len(units)),
                        text=source_text,
                        context=self._row_context(row),
                        metadata={
                            "adapter": self.name,
                            "file_key": file_key,
                            "row_index": row_index,
                        },
                    )
                )
        if not units:
            raise ValueError(f"Unsupported Translator++ project file: {relative_path}")
        return units

    def apply_translations(self, relative_path: str, text: str, translations: Mapping[str, str]) -> str:
        root = self._load_root(text)
        if root is None:
            raise ValueError(f"Unsupported Translator++ project file: {relative_path}")

        unit_index = 0
        for _file_key, data_rows in self._iter_data_rows(root):
            for _row_index, row in data_rows:
                source_text = self._row_source_text(row)
                if source_text is None or not source_text.strip():
                    continue
                translation = translations.get(str(unit_index))
                if translation is not None:
                    self._patch_row_translation(row, _decode_compressed_line_stream(translation))
                unit_index += 1
        return json.dumps(root, ensure_ascii=False, indent=2) + "\n"

    @staticmethod
    def _load_root(text: str) -> dict[str, object] | None:
        try:
            loaded: object = json.loads(text)
        except json.JSONDecodeError:
            return None
        if not isinstance(loaded, dict):
            return None
        root = cast("dict[str, object]", loaded)
        return root if TranslatorPlusPlusTransAdapter._files(root) is not None else None

    @staticmethod
    def _files(root: Mapping[str, object]) -> Mapping[str, object] | None:
        project = root.get("project")
        if isinstance(project, dict):
            files = cast("Mapping[str, object]", project).get("files")
            if isinstance(files, dict):
                return cast("Mapping[str, object]", files)
        files = root.get("files")
        if isinstance(files, dict):
            return cast("Mapping[str, object]", files)
        return None

    @classmethod
    def _iter_data_rows(cls, root: Mapping[str, object]) -> Iterable[tuple[str, list[tuple[int, object]]]]:
        files = cls._files(root)
        if files is None:
            return []

        rows_by_file: list[tuple[str, list[tuple[int, object]]]] = []
        for file_key, file_payload in files.items():
            if not isinstance(file_payload, dict):
                continue
            data = cast("Mapping[str, object]", file_payload).get("data")
            if not isinstance(data, list):
                continue
            rows_by_file.append((str(file_key), list(enumerate(data))))
        return rows_by_file

    @staticmethod
    def _row_source_text(row: object) -> str | None:
        if isinstance(row, list):
            return row[0] if row and isinstance(row[0], str) else None
        if not isinstance(row, dict):
            return None
        row_mapping = cast("Mapping[str, object]", row)
        for key in ("original", "source", "text"):
            value = row_mapping.get(key)
            if isinstance(value, str):
                return value
        return None

    @staticmethod
    def _row_context(row: object) -> str | None:
        if not isinstance(row, dict):
            return None
        row_mapping = cast("Mapping[str, object]", row)
        for key in ("rowInfoText", "context", "tags"):
            value = row_mapping.get(key)
            if isinstance(value, str) and value.strip():
                return value
        return None

    @staticmethod
    def _patch_row_translation(row: object, translation: str) -> None:
        if isinstance(row, list):
            if len(row) >= 2:
                row[1] = translation
            else:
                row.append(translation)
            return
        if not isinstance(row, dict):
            return
        row_mapping = cast("dict[str, object]", row)
        for key in ("translation", "target", "translated", "machineTrans"):
            if key in row_mapping:
                row_mapping[key] = translation
                return
        row_mapping["translation"] = translation


class VnTextJsonAdapter:
    """Adapter for VNText JSON lists containing message/name entries."""

    name = "vntext_json"
    mime_type = "application/x-contextweave-galgame-vntext-json"
    supported_extensions = frozenset({".json"})
    stores_binary_as_base64 = False

    def can_read_file(self, path: Path) -> bool:
        if path.suffix.lower() not in self.supported_extensions or not path.is_file():
            return False
        try:
            text = path.read_text(encoding="utf-8-sig")
        except UnicodeDecodeError:
            return False
        return self._load_entries(text) is not None

    def extract_units(self, relative_path: str, text: str) -> list[TranslationUnit]:
        entries = self._load_entries(text)
        if entries is None:
            raise ValueError(f"Unsupported VNText JSON file: {relative_path}")

        units: list[TranslationUnit] = []
        for entry_index, entry in entries:
            message = cast("str", entry["message"])
            units.append(
                TranslationUnit(
                    relative_path=relative_path,
                    unit_id=str(len(units)),
                    text=message,
                    speaker=self._speaker(entry),
                    metadata={"adapter": self.name, "entry_index": entry_index},
                )
            )
        return units

    def apply_translations(self, relative_path: str, text: str, translations: Mapping[str, str]) -> str:
        loaded = self._load_list(text)
        if loaded is None:
            raise ValueError(f"Unsupported VNText JSON file: {relative_path}")

        unit_index = 0
        for item in loaded:
            if not isinstance(item, dict):
                continue
            entry = cast("dict[str, object]", item)
            message = entry.get("message")
            if not isinstance(message, str) or not message.strip():
                continue
            translation = translations.get(str(unit_index))
            if translation is not None:
                entry["message"] = _decode_compressed_line_stream(translation)
            unit_index += 1
        return json.dumps(loaded, ensure_ascii=False, indent=2) + "\n"

    @classmethod
    def _load_entries(cls, text: str) -> list[tuple[int, dict[str, object]]] | None:
        loaded = cls._load_list(text)
        if loaded is None:
            return None
        entries: list[tuple[int, dict[str, object]]] = []
        for index, item in enumerate(loaded):
            if not isinstance(item, dict):
                return None
            entry = cast("dict[str, object]", item)
            message = entry.get("message")
            if isinstance(message, str) and message.strip():
                entries.append((index, entry))
        return entries or None

    @staticmethod
    def _load_list(text: str) -> list[object] | None:
        try:
            loaded: object = json.loads(text)
        except json.JSONDecodeError:
            return None
        if not isinstance(loaded, list):
            return None
        return cast("list[object]", loaded)

    @staticmethod
    def _speaker(entry: Mapping[str, object]) -> str | None:
        name = entry.get("name")
        if isinstance(name, str) and name.strip():
            return name
        names = entry.get("names")
        if isinstance(names, list):
            text_names = [name for name in names if isinstance(name, str) and name.strip()]
            if text_names:
                return " / ".join(text_names)
        return None


class ParaTranzJsonAdapter:
    """Adapter for ParaTranz JSON files containing original/translation entries."""

    name = "paratranz_json"
    mime_type = "application/x-contextweave-galgame-paratranz-json"
    supported_extensions = frozenset({".json"})
    stores_binary_as_base64 = False

    def can_read_file(self, path: Path) -> bool:
        if path.suffix.lower() not in self.supported_extensions or not path.is_file():
            return False
        try:
            text = path.read_text(encoding="utf-8-sig")
        except UnicodeDecodeError:
            return False
        return self._load_entries(text) is not None

    def extract_units(self, relative_path: str, text: str) -> list[TranslationUnit]:
        entries = self._load_entries(text)
        if entries is None:
            raise ValueError(f"Unsupported ParaTranz JSON file: {relative_path}")

        units: list[TranslationUnit] = []
        for entry_index, entry in entries:
            original = cast("str", entry["original"])
            units.append(
                TranslationUnit(
                    relative_path=relative_path,
                    unit_id=str(len(units)),
                    text=original,
                    speaker=_optional_entry_str(entry, "speaker") or _optional_entry_str(entry, "name"),
                    context=_optional_entry_str(entry, "context"),
                    metadata={"adapter": self.name, "entry_index": entry_index},
                )
            )
        return units

    def apply_translations(self, relative_path: str, text: str, translations: Mapping[str, str]) -> str:
        loaded = self._load_list(text)
        if loaded is None:
            raise ValueError(f"Unsupported ParaTranz JSON file: {relative_path}")

        unit_index = 0
        for item in loaded:
            if not isinstance(item, dict):
                continue
            entry = cast("dict[str, object]", item)
            original = entry.get("original")
            if not isinstance(original, str) or not original.strip():
                continue
            translation = translations.get(str(unit_index))
            if translation is not None:
                entry[self._translation_key(entry)] = _decode_compressed_line_stream(translation)
            unit_index += 1
        return json.dumps(loaded, ensure_ascii=False, indent=2) + "\n"

    @classmethod
    def _load_entries(cls, text: str) -> list[tuple[int, dict[str, object]]] | None:
        loaded = cls._load_list(text)
        if loaded is None:
            return None
        entries: list[tuple[int, dict[str, object]]] = []
        for index, item in enumerate(loaded):
            if not isinstance(item, dict):
                return None
            entry = cast("dict[str, object]", item)
            original = entry.get("original")
            if isinstance(original, str) and original.strip():
                entries.append((index, entry))
        return entries or None

    @staticmethod
    def _load_list(text: str) -> list[object] | None:
        try:
            loaded: object = json.loads(text)
        except json.JSONDecodeError:
            return None
        if not isinstance(loaded, list):
            return None
        return cast("list[object]", loaded)

    @staticmethod
    def _translation_key(entry: Mapping[str, object]) -> str:
        for key in ("translation", "trans", "translated", "machineTrans"):
            if key in entry:
                return key
        return "translation"


class RenPyScriptAdapter:
    """Adapter for directly editable Ren'Py `.rpy` script files."""

    name = "renpy_rpy"
    mime_type = "application/x-contextweave-galgame-renpy-rpy"
    supported_extensions = frozenset({".rpy"})
    stores_binary_as_base64 = False

    def can_read_file(self, path: Path) -> bool:
        if path.suffix.lower() not in self.supported_extensions or not path.is_file():
            return False
        try:
            text = path.read_text(encoding="utf-8-sig")
        except UnicodeDecodeError:
            return False
        return bool(self.extract_units(path.name, text))

    def extract_units(self, relative_path: str, text: str) -> list[TranslationUnit]:
        lines = text.splitlines(keepends=True)
        units: list[TranslationUnit] = []
        for line_index, line in enumerate(lines):
            parsed = _parse_renpy_translatable_line(line)
            if parsed is None or parsed.kind == "new":
                continue

            target_line_index = line_index
            if parsed.kind == "old":
                new_line_index = _find_following_renpy_new_line(lines, line_index)
                if new_line_index is None:
                    continue
                target_line_index = new_line_index

            units.append(
                TranslationUnit(
                    relative_path=relative_path,
                    unit_id=str(len(units)),
                    text=parsed.text,
                    speaker=parsed.speaker,
                    metadata={
                        "adapter": self.name,
                        "kind": parsed.kind,
                        "line": line_index + 1,
                        "target_line": target_line_index + 1,
                    },
                )
            )
        return units

    def apply_translations(self, relative_path: str, text: str, translations: Mapping[str, str]) -> str:
        lines = text.splitlines(keepends=True)
        units = self.extract_units(relative_path, text)
        for unit in units:
            translation = translations.get(unit.unit_id)
            if translation is None:
                continue
            translation = _decode_compressed_line_stream(translation)
            target_line = _metadata_int(unit.metadata, "target_line")
            if target_line is None or target_line < 1 or target_line > len(lines):
                raise ValueError(f"Invalid Ren'Py target line for unit {unit.unit_id} in {relative_path}")
            lines[target_line - 1] = _replace_first_quoted_string(lines[target_line - 1], translation)
        return "".join(lines)


class KagKsScriptAdapter:
    """Adapter for TyranoScript/KAG/KiriKiri extracted `.ks` text scripts."""

    name = "kag_ks"
    mime_type = "application/x-contextweave-galgame-kag-ks"
    supported_extensions = frozenset({".ks"})
    stores_binary_as_base64 = False

    def can_read_file(self, path: Path) -> bool:
        if path.suffix.lower() not in self.supported_extensions or not path.is_file():
            return False
        try:
            text = path.read_text(encoding="utf-8-sig")
        except UnicodeDecodeError:
            return False
        return bool(self.extract_units(path.name, text))

    def extract_units(self, relative_path: str, text: str) -> list[TranslationUnit]:
        units: list[TranslationUnit] = []
        speaker: str | None = None
        for line_index, line in enumerate(text.splitlines(keepends=True)):
            stripped_line = line.strip()
            if not stripped_line:
                continue
            if stripped_line.startswith("#") and len(stripped_line) > 1:
                speaker = stripped_line[1:].strip() or None
                continue
            if not _is_kag_dialogue_line(stripped_line):
                continue
            units.append(
                TranslationUnit(
                    relative_path=relative_path,
                    unit_id=str(len(units)),
                    text=stripped_line,
                    speaker=speaker,
                    metadata={"adapter": self.name, "line": line_index + 1},
                )
            )
        return units

    def apply_translations(self, relative_path: str, text: str, translations: Mapping[str, str]) -> str:
        lines = text.splitlines(keepends=True)
        units = self.extract_units(relative_path, text)
        for unit in units:
            translation = translations.get(unit.unit_id)
            if translation is None:
                continue
            translation = _decode_compressed_line_stream(translation)
            line_number = _metadata_int(unit.metadata, "line")
            if line_number is None or line_number < 1 or line_number > len(lines):
                raise ValueError(f"Invalid KAG target line for unit {unit.unit_id} in {relative_path}")
            lines[line_number - 1] = _replace_kag_line_text(lines[line_number - 1], translation)
        return "".join(lines)


class RpgMakerMvMzJsonAdapter:
    """Adapter for RPG Maker MV/MZ extracted `data/*.json` event command files."""

    name = "rpg_maker_mv_mz_json"
    mime_type = "application/x-contextweave-galgame-rpg-maker-mv-mz-json"
    supported_extensions = frozenset({".json"})
    stores_binary_as_base64 = False

    def can_read_file(self, path: Path) -> bool:
        if path.suffix.lower() not in self.supported_extensions or not path.is_file():
            return False
        try:
            text = path.read_text(encoding="utf-8-sig")
        except UnicodeDecodeError:
            return False
        try:
            return bool(self.extract_units(path.name, text))
        except ValueError:
            return False

    def extract_units(self, relative_path: str, text: str) -> list[TranslationUnit]:
        root, kind = self._load_supported_root(relative_path, text)
        slots = _rpg_maker_text_slots(kind, root)
        if not slots:
            raise ValueError(f"Unsupported RPG Maker MV/MZ JSON file: {relative_path}")

        units: list[TranslationUnit] = []
        for slot in slots:
            metadata = dict(slot.metadata)
            metadata["adapter"] = self.name
            units.append(
                TranslationUnit(
                    relative_path=relative_path,
                    unit_id=str(len(units)),
                    text=slot.text,
                    speaker=slot.speaker,
                    context=slot.context,
                    metadata=metadata,
                )
            )
        return units

    def apply_translations(self, relative_path: str, text: str, translations: Mapping[str, str]) -> str:
        root, kind = self._load_supported_root(relative_path, text)
        slots = _rpg_maker_text_slots(kind, root)
        if not slots:
            raise ValueError(f"Unsupported RPG Maker MV/MZ JSON file: {relative_path}")

        for slot_index, slot in enumerate(slots):
            translation = translations.get(str(slot_index))
            if translation is None:
                continue
            _patch_rpg_maker_slot(slot, _decode_compressed_line_stream(translation))
        return json.dumps(root, ensure_ascii=False, indent=2) + "\n"

    @staticmethod
    def _load_supported_root(relative_path: str, text: str) -> tuple[object, str]:
        kind = _rpg_maker_file_kind(relative_path)
        if kind is None:
            raise ValueError(f"Unsupported RPG Maker MV/MZ JSON file: {relative_path}")
        try:
            root = json.loads(text)
        except json.JSONDecodeError as exc:
            raise ValueError(f"Unsupported RPG Maker MV/MZ JSON file: {relative_path}") from exc
        return root, kind


class _SpreadsheetXlsxAdapter:
    """Base adapter for extracted galgame spreadsheet workbooks."""

    name: str
    mime_type: str
    supported_extensions = frozenset({".xlsx"})
    stores_binary_as_base64 = True
    source_headers: tuple[str, ...]
    translation_headers: tuple[str, ...]
    speaker_headers = ("speaker", "name", "character", "角色", "名称", "名字")
    context_headers = ("context", "rowinfotext", "tags", "note", "notes", "备注", "上下文", "注释")

    def can_read_file(self, path: Path) -> bool:
        if path.suffix.lower() not in self.supported_extensions or not path.is_file():
            return False
        try:
            source_text = _encode_binary_source(path.read_bytes())
            return bool(self.extract_units(path.name, source_text))
        except (BadZipFile, InvalidFileException, OSError, ValueError):
            return False

    def extract_units(self, relative_path: str, text: str) -> list[TranslationUnit]:
        workbook = self._load_workbook(relative_path, text)
        units: list[TranslationUnit] = []
        for sheet in workbook.worksheets:
            columns = self._find_columns(sheet)
            if columns is None:
                continue
            for row_number in range(columns.header_row + 1, int(sheet.max_row) + 1):
                source_text = _spreadsheet_source_text(sheet.cell(row=row_number, column=columns.source_column).value)
                if source_text is None:
                    continue
                units.append(
                    TranslationUnit(
                        relative_path=relative_path,
                        unit_id=str(len(units)),
                        text=source_text,
                        speaker=self._optional_cell_text(sheet, row_number, columns.speaker_column),
                        context=self._optional_cell_text(sheet, row_number, columns.context_column),
                        metadata={
                            "adapter": self.name,
                            "sheet": str(sheet.title),
                            "row": row_number,
                            "source_column": columns.source_column,
                            "translation_column": columns.translation_column,
                        },
                    )
                )
        if not units:
            raise ValueError(f"Unsupported galgame spreadsheet file: {relative_path}")
        return units

    def apply_translations(self, relative_path: str, text: str, translations: Mapping[str, str]) -> str:
        workbook = self._load_workbook(relative_path, text)
        unit_index = 0
        for sheet in workbook.worksheets:
            columns = self._find_columns(sheet)
            if columns is None:
                continue
            translation_column = columns.translation_column or self._append_translation_column(sheet, columns.header_row)
            for row_number in range(columns.header_row + 1, int(sheet.max_row) + 1):
                source_text = _spreadsheet_source_text(sheet.cell(row=row_number, column=columns.source_column).value)
                if source_text is None:
                    continue
                translation = translations.get(str(unit_index))
                if translation is not None:
                    sheet.cell(row=row_number, column=translation_column).value = _decode_compressed_line_stream(
                        translation
                    )
                unit_index += 1

        output = BytesIO()
        workbook.save(output)
        return _encode_binary_source(output.getvalue())

    @staticmethod
    def _load_workbook(relative_path: str, text: str) -> Any:
        try:
            return load_workbook(filename=BytesIO(_decode_binary_source(text)))
        except (BadZipFile, InvalidFileException, OSError, ValueError) as exc:
            raise ValueError(f"Unsupported galgame spreadsheet file: {relative_path}") from exc

    def _find_columns(self, sheet: Any) -> _SpreadsheetColumns | None:
        max_header_row = min(int(sheet.max_row), 5)
        max_column = int(sheet.max_column)
        for header_row in range(1, max_header_row + 1):
            header_map = self._header_map(sheet, header_row, max_column)
            source_column = _first_matching_header(header_map, self.source_headers)
            if source_column is None:
                continue
            return _SpreadsheetColumns(
                header_row=header_row,
                source_column=source_column,
                translation_column=_first_matching_header(header_map, self.translation_headers),
                speaker_column=_first_matching_header(header_map, self.speaker_headers),
                context_column=_first_matching_header(header_map, self.context_headers),
            )
        return None

    @staticmethod
    def _header_map(sheet: Any, header_row: int, max_column: int) -> dict[str, int]:
        headers: dict[str, int] = {}
        for column in range(1, max_column + 1):
            value = sheet.cell(row=header_row, column=column).value
            normalized = _normalize_spreadsheet_header(value)
            if normalized:
                headers.setdefault(normalized, column)
        return headers

    @staticmethod
    def _optional_cell_text(sheet: Any, row_number: int, column: int | None) -> str | None:
        if column is None:
            return None
        return _spreadsheet_optional_text(sheet.cell(row=row_number, column=column).value)

    @staticmethod
    def _append_translation_column(sheet: Any, header_row: int) -> int:
        column = int(sheet.max_column) + 1
        sheet.cell(row=header_row, column=column).value = "Translation"
        return column


class TppXlsxAdapter(_SpreadsheetXlsxAdapter):
    """Adapter for TPP / Translator++ spreadsheet exports."""

    name = "tpp_xlsx"
    mime_type = "application/x-contextweave-galgame-tpp-xlsx"
    source_headers = ("originaltext", "original", "sourcetext", "source")
    translation_headers = ("initial", "translation", "translatedtext", "target", "targettext")


class WolfRpgXlsxAdapter(_SpreadsheetXlsxAdapter):
    """Adapter for Wolf RPG extracted spreadsheets with Chinese/Japanese-style headers."""

    name = "wolf_rpg_xlsx"
    mime_type = "application/x-contextweave-galgame-wolf-rpg-xlsx"
    source_headers = ("原文", "原始文本", "源文本", "文本", "source", "original")
    translation_headers = ("译文", "翻译", "译文文本", "目标文本", "translation", "target")


class GalgameDocument(Document):
    """Document for offline galgame script/intermediate-file translation."""

    document_type = "galgame"
    supported_export_formats = ("native",)
    requires_ocr_config = False
    ocr_required_for_translation = False
    supports_preserve_structure = True
    supports_multi_export = False
    supports_original_image_export = False

    def __init__(self, repo: DocumentRepository, document_id: int):
        super().__init__(repo, document_id)
        self._translated_lines: list[str] | None = None

    @classmethod
    def can_import(cls, path: Path) -> bool:
        path = Path(path)
        if not path.exists():
            return False
        return bool(_scan_importable_files(path))

    @classmethod
    def can_import_with_adapter(cls, path: Path, adapter_name: str) -> bool:
        path = Path(path)
        if not path.exists():
            return False
        return bool(_scan_importable_files(path, adapter_name=adapter_name))

    @classmethod
    def do_import(
        cls,
        repo: DocumentRepository,
        path: Path,
        cancel_check: Callable[[], bool] | None = None,
        adapter_name: str | None = None,
    ) -> dict[str, int]:
        path = Path(path)
        raise_if_cancelled(cancel_check)
        base_path = path.parent if path.is_file() else path

        files_to_import: list[tuple[Path, str, str, GalgameAdapter]] = []
        skipped = 0
        for file_path, adapter in _scan_importable_files(path, adapter_name=adapter_name):
            raise_if_cancelled(cancel_check)
            text_content = _read_source_for_adapter(file_path, adapter)
            relative_path = file_path.relative_to(base_path).as_posix()
            if not adapter.extract_units(relative_path, text_content):
                skipped += 1
                continue
            if repo.source_exists_by_content(text_content):
                skipped += 1
                continue
            files_to_import.append((file_path, relative_path, text_content, adapter))

        if not files_to_import:
            return {"imported": 0, "skipped": skipped}

        imported = 0
        repo.begin()
        try:
            raise_if_cancelled(cancel_check)
            document_id = repo.insert_document(cls.document_type, auto_commit=False)
            for sequence_number, (_file_path, relative_path, text_content, adapter) in enumerate(files_to_import):
                raise_if_cancelled(cancel_check)
                repo.insert_document_source(
                    document_id,
                    sequence_number,
                    "text",
                    relative_path=relative_path,
                    text_content=text_content,
                    mime_type=adapter.mime_type,
                    is_ocr_completed=True,
                    is_text_added=False,
                    auto_commit=False,
                )
                imported += 1
            raise_if_cancelled(cancel_check)
            repo.commit()
        except Exception:
            repo.rollback()
            raise

        return {"imported": imported, "skipped": skipped}

    def is_ocr_completed(self) -> bool:
        return True

    async def process_ocr(
        self,
        llm_client: LLMClient,
        source_ids: list[int] | None = None,
        cancel_check: Callable[[], bool] | None = None,
        on_item_processed: Callable[[], None] | None = None,
    ) -> int:
        _ = (llm_client, source_ids, on_item_processed)
        raise_if_cancelled(cancel_check)
        return 0

    def get_text(self) -> str:
        return "\n".join(unit.text for unit in self._all_translation_units())

    def get_glossary_seeds(self) -> tuple[GalgameGlossarySeed, ...]:
        seeds: list[GalgameGlossarySeed] = []
        for source in self._ordered_sources():
            relative_path, text_content, adapter = _source_adapter(source)
            units = adapter.extract_units(relative_path, text_content)
            seeds.extend(_glossary_seeds_from_units(units))
            seeds.extend(_source_metadata_glossary_seeds(relative_path, text_content))
        return _dedupe_glossary_seeds(seeds)

    def is_text_added(self) -> bool:
        sources = self.repo.get_document_sources(self.document_id)
        if not sources:
            return True
        return all(source["is_text_added"] == 1 for source in sources)

    def mark_text_added(self) -> None:
        self.repo.update_all_sources_text_added(self.document_id)

    async def set_text(
        self,
        lines: list[str],
        cancel_check: Callable[[], bool] | None = None,
        progress_callback: ProgressCallback | None = None,
    ) -> int:
        _ = progress_callback
        raise_if_cancelled(cancel_check)
        self._translated_lines = _expand_galgame_export_lines(lines, self._all_translation_units())
        return len(lines)

    async def reembed(
        self,
        image_reembedding_config: ImageReembeddingConfig,
        *,
        force: bool = False,
        source_ids: list[int] | None = None,
        cancel_check: Callable[[], bool] | None = None,
        progress_callback: ProgressCallback | None = None,
    ) -> int:
        _ = (image_reembedding_config, force, source_ids, progress_callback)
        raise_if_cancelled(cancel_check)
        return 0

    def can_export(self, export_format: str) -> bool:
        normalized = export_format.lower().lstrip(".")
        return normalized == "native"

    @classmethod
    def export_merged(
        cls,
        documents: list[Document],
        export_format: str,
        output_path: Path,
        *,
        use_original_images: bool = False,
    ) -> None:
        _ = (documents, export_format, output_path, use_original_images)
        raise ValueError(
            "Galgame documents must be exported with preserve-structure export so each imported source "
            "is patched with its original adapter."
        )

    def export_preserve_structure(self, output_folder: Path) -> None:
        output_folder = Path(output_folder)
        for source, adapter, _units, translations in self._source_translation_batches():
            relative_path, text_content, _adapter = _source_adapter(source)
            output_path = _safe_output_path(output_folder, relative_path)
            output_path.parent.mkdir(parents=True, exist_ok=True)
            patched_text = adapter.apply_translations(relative_path, text_content, translations)
            if adapter.stores_binary_as_base64:
                output_path.write_bytes(_decode_binary_source(patched_text))
            else:
                output_path.write_text(patched_text, encoding="utf-8")

    def _ordered_sources(self) -> list[dict[str, object]]:
        sources = cast("list[dict[str, object]]", self.repo.get_document_sources(self.document_id))
        return sorted(sources, key=_source_sequence_number)

    def _all_translation_units(self) -> list[TranslationUnit]:
        units: list[TranslationUnit] = []
        for source in self._ordered_sources():
            relative_path, text_content, adapter = _source_adapter(source)
            units.extend(adapter.extract_units(relative_path, text_content))
        return units

    def _source_translation_batches(
        self,
    ) -> list[tuple[dict[str, object], GalgameAdapter, list[TranslationUnit], dict[str, str]]]:
        if self._translated_lines is None:
            raise ValueError("No translated text to export. Call set_text() first.")

        cursor = 0
        batches: list[tuple[dict[str, object], GalgameAdapter, list[TranslationUnit], dict[str, str]]] = []
        for source in self._ordered_sources():
            relative_path, text_content, adapter = _source_adapter(source)
            units = adapter.extract_units(relative_path, text_content)
            if cursor + len(units) > len(self._translated_lines):
                raise ValueError(
                    "Translated galgame unit count is shorter than the source unit stream "
                    f"for document {self.document_id}."
                )
            translations = {
                unit.unit_id: self._translated_lines[cursor + index] for index, unit in enumerate(units)
            }
            issues = validate_galgame_translations(units, translations)
            if issues:
                issue = issues[0]
                raise ValueError(
                    f"Invalid galgame translation for {issue.relative_path} unit {issue.unit_id}: {issue.message}"
                )
            cursor += len(units)
            batches.append((source, adapter, units, translations))

        extra_lines = [line for line in self._translated_lines[cursor:] if line]
        if extra_lines:
            raise ValueError(
                "Translated galgame unit count is longer than the source unit stream "
                f"for document {self.document_id}."
            )
        return batches


def get_galgame_adapters() -> tuple[GalgameAdapter, ...]:
    return (
        TranslatorPlusPlusTransAdapter(),
        TppXlsxAdapter(),
        WolfRpgXlsxAdapter(),
        RenPyScriptAdapter(),
        KagKsScriptAdapter(),
        RpgMakerMvMzJsonAdapter(),
        VnTextJsonAdapter(),
        ParaTranzJsonAdapter(),
        MToolJsonAdapter(),
    )


def get_galgame_adapter(adapter_name: str) -> GalgameAdapter:
    for adapter in get_galgame_adapters():
        if adapter.name == adapter_name:
            return adapter
    raise ValueError(f"Unknown galgame adapter: {adapter_name}")


def get_galgame_bridge_formats() -> tuple[GalgameBridgeFormat, ...]:
    return _BRIDGE_FORMATS


def get_galgame_external_tool_specs() -> tuple[GalgameExternalToolSpec, ...]:
    return _EXTERNAL_TOOL_SPECS


def inspect_galgame_external_helpers(
    path: Path,
    *,
    output_folder: Path | None = None,
    tool_names: Iterable[str] | None = None,
    executable_resolver: Callable[[str], str | None] = shutil.which,
) -> GalgameExternalHelperSummary:
    path = Path(path)
    if not path.exists():
        return GalgameExternalHelperSummary(
            commands=(),
            issues=(GalgameExternalHelperIssue(str(path), "", "path does not exist"),),
        )

    base_path = path.parent if path.is_file() else path
    base_output = output_folder or _default_external_output_folder(path)
    allowed_tools = set(tool_names) if tool_names is not None else None
    commands: list[GalgameExternalHelperCommand] = []
    issues: list[GalgameExternalHelperIssue] = []
    for file_path in _external_helper_candidates(path):
        relative_path = file_path.relative_to(base_path).as_posix()
        extension = file_path.suffix.lower()
        specs = _external_specs_for_extension(extension, allowed_tools)
        if not specs:
            issues.append(
                GalgameExternalHelperIssue(
                    relative_path,
                    extension,
                    "external extraction is known but no safe helper wrapper is configured",
                )
            )
            continue

        command = _first_available_external_command(file_path, relative_path, base_path, base_output, specs, executable_resolver)
        if command is None:
            tool_list = ", ".join(spec.name for spec in specs)
            issues.append(
                GalgameExternalHelperIssue(
                    relative_path,
                    extension,
                    f"install one of these external helpers and run explicitly: {tool_list}",
                )
            )
            continue
        commands.append(command)
    return GalgameExternalHelperSummary(commands=tuple(commands), issues=tuple(issues))


def inspect_galgame_import(path: Path, adapter_name: str | None = None) -> GalgameImportSummary:
    path = Path(path)
    if not path.exists():
        return GalgameImportSummary(files=())
    base_path = path.parent if path.is_file() else path
    files: list[GalgameImportSourceSummary] = []
    skipped = 0
    for file_path, adapter in _scan_importable_files(path, adapter_name=adapter_name):
        relative_path = file_path.relative_to(base_path).as_posix()
        try:
            text_content = _read_source_for_adapter(file_path, adapter)
            unit_count = len(adapter.extract_units(relative_path, text_content))
        except (OSError, UnicodeDecodeError, ValueError):
            skipped += 1
            continue
        files.append(
            GalgameImportSourceSummary(
                relative_path=relative_path,
                adapter_name=adapter.name,
                unit_count=unit_count,
                confidence=_adapter_detection_confidence(adapter, relative_path),
            )
        )
    return GalgameImportSummary(files=tuple(files), skipped=skipped)


def inspect_galgame_glossary_seeds(path: Path, adapter_name: str | None = None) -> tuple[GalgameGlossarySeed, ...]:
    path = Path(path)
    if not path.exists():
        return ()
    base_path = path.parent if path.is_file() else path
    seeds: list[GalgameGlossarySeed] = []
    for file_path in _source_candidate_files(path):
        relative_path = file_path.relative_to(base_path).as_posix()
        seeds.extend(_metadata_glossary_seeds_from_file(file_path, relative_path))
        for adapter in _selected_galgame_adapters(adapter_name):
            if file_path.suffix.lower() not in adapter.supported_extensions:
                continue
            try:
                text_content = _read_source_for_adapter(file_path, adapter)
                units = adapter.extract_units(relative_path, text_content)
            except (OSError, UnicodeDecodeError, ValueError):
                continue
            seeds.extend(_glossary_seeds_from_units(units))
            break
    return _dedupe_glossary_seeds(seeds)


def validate_galgame_translations(
    units: Iterable[TranslationUnit], translations: Mapping[str, str]
) -> list[GalgameValidationIssue]:
    issues: list[GalgameValidationIssue] = []
    for unit in units:
        translation = translations.get(unit.unit_id)
        if translation is None or not translation.strip():
            issues.append(GalgameValidationIssue(unit.relative_path, unit.unit_id, "missing translation"))
            continue
        issue = _translation_validation_message(unit, translation)
        if issue is not None:
            issues.append(GalgameValidationIssue(unit.relative_path, unit.unit_id, issue))
    return issues


def _encode_binary_source(content: bytes) -> str:
    return base64.b64encode(content).decode("ascii")


def _decode_binary_source(text: str) -> bytes:
    try:
        return base64.b64decode(text.encode("ascii"), validate=True)
    except (binascii.Error, UnicodeEncodeError) as exc:
        raise ValueError("Invalid encoded galgame binary source.") from exc


def _default_external_output_folder(path: Path) -> Path:
    if path.is_file():
        return path.parent / f"{path.stem}_extracted"
    return path / "extracted"


def _source_candidate_files(path: Path) -> list[Path]:
    return [path] if path.is_file() else sorted(file_path for file_path in path.rglob("*") if file_path.is_file())


def _external_helper_candidates(path: Path) -> list[Path]:
    return [file_path for file_path in _source_candidate_files(path) if file_path.suffix.lower() in _EXTERNAL_ARCHIVE_EXTENSIONS]


def _external_specs_for_extension(
    extension: str,
    allowed_tools: set[str] | None,
) -> tuple[GalgameExternalToolSpec, ...]:
    return tuple(
        spec
        for spec in _EXTERNAL_TOOL_SPECS
        if extension in spec.supported_extensions and (allowed_tools is None or spec.name in allowed_tools)
    )


def _first_available_external_command(
    file_path: Path,
    relative_path: str,
    base_path: Path,
    base_output: Path,
    specs: Iterable[GalgameExternalToolSpec],
    executable_resolver: Callable[[str], str | None],
) -> GalgameExternalHelperCommand | None:
    for spec in specs:
        executable = _resolve_external_executable(spec, executable_resolver)
        if executable is None:
            continue
        output_folder = _external_output_folder_for_file(base_output, relative_path)
        return GalgameExternalHelperCommand(
            source_path=relative_path,
            output_folder=output_folder.relative_to(base_path).as_posix()
            if output_folder.is_relative_to(base_path)
            else output_folder.as_posix(),
            tool_name=spec.name,
            executable=executable,
            argv=_external_helper_argv(spec, executable, file_path, output_folder),
            note=spec.note,
        )
    return None


def _resolve_external_executable(
    spec: GalgameExternalToolSpec,
    executable_resolver: Callable[[str], str | None],
) -> str | None:
    for executable_name in spec.executable_names:
        executable = executable_resolver(executable_name)
        if executable:
            return executable
    return None


def _external_output_folder_for_file(base_output: Path, relative_path: str) -> Path:
    relative = PurePosixPath(relative_path).with_suffix("")
    return base_output.joinpath(*relative.parts)


def _external_helper_argv(
    spec: GalgameExternalToolSpec,
    executable: str,
    file_path: Path,
    output_folder: Path,
) -> tuple[str, ...]:
    source = file_path.as_posix()
    output = output_folder.as_posix()
    if spec.name == "rpatool":
        return (executable, "-x", source, "-o", output)
    if spec.name == "unrpyc":
        return (executable, "--clobber", "-o", output, source)
    if spec.name == "garbro":
        return (executable, "x", source, output)
    return (executable, source, output)


def _scan_importable_files(path: Path, adapter_name: str | None = None) -> list[tuple[Path, GalgameAdapter]]:
    candidates = [path] if path.is_file() else sorted(file_path for file_path in path.rglob("*") if file_path.is_file())
    base_path = path.parent if path.is_file() else path
    adapters = _selected_galgame_adapters(adapter_name)

    matches: list[tuple[Path, GalgameAdapter]] = []
    for file_path in candidates:
        for adapter in adapters:
            if file_path.suffix.lower() not in adapter.supported_extensions:
                continue
            relative_path = file_path.relative_to(base_path).as_posix()
            try:
                text_content = _read_source_for_adapter(file_path, adapter)
                if adapter.extract_units(relative_path, text_content):
                    matches.append((file_path, adapter))
                    break
            except (OSError, UnicodeDecodeError, ValueError):
                continue
    return matches


def _selected_galgame_adapters(adapter_name: str | None) -> tuple[GalgameAdapter, ...]:
    if adapter_name is None:
        return get_galgame_adapters()
    return (get_galgame_adapter(adapter_name),)


def _read_source_for_adapter(file_path: Path, adapter: GalgameAdapter) -> str:
    if adapter.stores_binary_as_base64:
        return _encode_binary_source(file_path.read_bytes())
    return file_path.read_text(encoding="utf-8-sig")


def _adapter_detection_confidence(adapter: GalgameAdapter, relative_path: str) -> str:
    if adapter.name in {"vntext_json", "paratranz_json", "mtool_json"}:
        return "medium"
    if adapter.name == "rpg_maker_mv_mz_json":
        parts = [part.lower() for part in PurePosixPath(relative_path).parts]
        return "high" if "data" in parts else "medium"
    return "high"


def _source_adapter(source: Mapping[str, object]) -> tuple[str, str, GalgameAdapter]:
    relative_path_value = source.get("relative_path")
    relative_path = str(relative_path_value or "").strip()
    if not relative_path:
        raise ValueError("Galgame source does not have an original relative path.")
    text_content = str(source.get("text_content") or "")
    mime_type = str(source.get("mime_type") or "")
    adapter = next((candidate for candidate in get_galgame_adapters() if candidate.mime_type == mime_type), None)
    if adapter is None:
        raise ValueError(f"Unknown galgame adapter MIME type for {relative_path}: {mime_type}")
    if PurePosixPath(relative_path).suffix.lower() not in adapter.supported_extensions:
        raise ValueError(f"Galgame adapter {adapter.name} does not support source path: {relative_path}")
    try:
        if adapter.extract_units(relative_path, text_content):
            return relative_path, text_content, adapter
    except ValueError as exc:
        raise ValueError(f"Galgame adapter {adapter.name} cannot read source: {relative_path}") from exc
    raise ValueError(f"Galgame adapter {adapter.name} found no translation units in source: {relative_path}")


def _source_sequence_number(source: Mapping[str, object]) -> int:
    value = source.get("sequence_number", 0)
    if isinstance(value, int):
        return value
    if isinstance(value, str):
        try:
            return int(value)
        except ValueError:
            return 0
    return 0


def _safe_output_path(output_folder: Path, relative_path: str) -> Path:
    relative = PurePosixPath(relative_path)
    if relative.is_absolute() or any(part in {"", ".", ".."} for part in relative.parts):
        raise ValueError(f"Unsafe galgame source relative path: {relative_path}")
    return output_folder.joinpath(*relative.parts)


def _expand_galgame_export_lines(lines: list[str], units: list[TranslationUnit]) -> list[str]:
    line_stream = list(lines)
    if any("\n" in line or "\r" in line for line in line_stream):
        raise ValueError("Translated galgame line stream entries cannot contain newline characters.")

    source_line_counts = [_galgame_unit_line_count(unit.text) for unit in units]
    if len(line_stream) == sum(source_line_counts):
        source_aligned_lines: list[str] = []
        cursor = 0
        for source_line_count in source_line_counts:
            next_cursor = cursor + source_line_count
            source_aligned_lines.append("\n".join(line_stream[cursor:next_cursor]))
            cursor = next_cursor
        return source_aligned_lines

    expected_line_count = sum(source_line_counts)
    if len(line_stream) < expected_line_count:
        raise ValueError("Translated galgame line stream is shorter than the source unit stream.")
    raise ValueError("Translated galgame line stream is longer than the source unit stream.")


def _galgame_unit_line_count(text: str) -> int:
    normalized = text.replace("\r\n", "\n").replace("\r", "\n")
    return max(1, len(normalized.split("\n")))


def _translation_validation_message(unit: TranslationUnit, translation: str) -> str | None:
    if "\0" in translation:
        return "translation contains a NUL byte"
    adapter_name = str(unit.metadata.get("adapter") or "")
    if adapter_name in _LINE_CONSTRAINED_ADAPTERS and any(character in translation for character in "\r\n"):
        return "line-constrained script translations cannot contain newline characters"
    return None


def _decode_compressed_line_stream(text: str) -> str:
    return "\n".join(decode_compressed_line(line) for line in text.split("\n"))


def _tsv_escape(text: str) -> str:
    return text.replace("\\", "\\\\").replace("\t", "\\t").replace("\r", "\\r").replace("\n", "\\n")


def _required_str(payload: Mapping[str, object], key: str) -> str:
    value = payload.get(key)
    if not isinstance(value, str):
        raise ValueError(f"Galgame translation unit field '{key}' must be a string.")
    return value


def _optional_str(payload: Mapping[str, object], key: str) -> str | None:
    value = payload.get(key)
    if value is None:
        return None
    if not isinstance(value, str):
        raise ValueError(f"Galgame translation unit field '{key}' must be a string when present.")
    return value


def _optional_entry_str(entry: Mapping[str, object], key: str) -> str | None:
    value = entry.get(key)
    return value if isinstance(value, str) and value.strip() else None


def _metadata_glossary_seeds_from_file(file_path: Path, relative_path: str) -> list[GalgameGlossarySeed]:
    if file_path.suffix.lower() not in {".json", ".rpy"}:
        return []
    try:
        text_content = file_path.read_text(encoding="utf-8-sig")
    except (OSError, UnicodeDecodeError):
        return []
    return _source_metadata_glossary_seeds(relative_path, text_content)


def _source_metadata_glossary_seeds(relative_path: str, text_content: str) -> list[GalgameGlossarySeed]:
    suffix = PurePosixPath(relative_path).suffix.lower()
    if suffix == ".rpy":
        return _renpy_character_definition_seeds(relative_path, text_content)
    if suffix == ".json":
        return _rpg_maker_metadata_glossary_seeds(relative_path, text_content)
    return []


def _glossary_seeds_from_units(units: Iterable[TranslationUnit]) -> list[GalgameGlossarySeed]:
    seeds: list[GalgameGlossarySeed] = []
    for unit in units:
        if unit.speaker is None or not _is_glossary_seed_text(unit.speaker):
            continue
        seeds.append(
            GalgameGlossarySeed(
                term=unit.speaker.strip(),
                category="speaker",
                source=unit.relative_path,
                context=unit.context,
            )
        )
    return seeds


def _renpy_character_definition_seeds(relative_path: str, text_content: str) -> list[GalgameGlossarySeed]:
    seeds: list[GalgameGlossarySeed] = []
    for line in text_content.splitlines():
        match = _RENPY_CHARACTER_DEF_RE.match(line)
        if match is None:
            continue
        quoted = _find_first_quoted_string(line[match.end() - 1 :])
        if quoted is None or not _is_glossary_seed_text(quoted.text):
            continue
        seeds.append(
            GalgameGlossarySeed(
                term=quoted.text.strip(),
                category="character",
                source=relative_path,
                context=f"Ren'Py Character {match.group('symbol')}",
            )
        )
    return seeds


def _rpg_maker_metadata_glossary_seeds(relative_path: str, text_content: str) -> list[GalgameGlossarySeed]:
    name = PurePosixPath(relative_path).name.lower()
    if name not in {"actors.json", "system.json"}:
        return []
    try:
        root: object = json.loads(text_content)
    except json.JSONDecodeError:
        return []
    if name == "actors.json":
        return _rpg_maker_actor_glossary_seeds(relative_path, root)
    if name == "system.json":
        return _rpg_maker_system_glossary_seeds(relative_path, root)
    return []


def _rpg_maker_actor_glossary_seeds(relative_path: str, root: object) -> list[GalgameGlossarySeed]:
    if not isinstance(root, list):
        return []
    seeds: list[GalgameGlossarySeed] = []
    for actor_index, actor in enumerate(root):
        if not isinstance(actor, dict):
            continue
        actor_mapping = cast("Mapping[str, object]", actor)
        context = f"RPG Maker actor {actor_mapping.get('id') or actor_index}"
        for key in ("name", "nickname"):
            value = _optional_entry_str(actor_mapping, key)
            if value is not None and _is_glossary_seed_text(value):
                seeds.append(GalgameGlossarySeed(value.strip(), "character", relative_path, context))
    return seeds


def _rpg_maker_system_glossary_seeds(relative_path: str, root: object) -> list[GalgameGlossarySeed]:
    if not isinstance(root, dict):
        return []
    root_mapping = cast("Mapping[str, object]", root)
    seeds: list[GalgameGlossarySeed] = []
    for key, category in (("gameTitle", "project"), ("currencyUnit", "term")):
        value = _optional_entry_str(root_mapping, key)
        if value is not None and _is_glossary_seed_text(value):
            seeds.append(GalgameGlossarySeed(value.strip(), category, relative_path, f"RPG Maker System.{key}"))
    return seeds


def _is_glossary_seed_text(text: str) -> bool:
    seed = text.strip()
    return bool(seed) and len(seed) <= 80 and "\0" not in seed


def _dedupe_glossary_seeds(seeds: Iterable[GalgameGlossarySeed]) -> tuple[GalgameGlossarySeed, ...]:
    deduped: list[GalgameGlossarySeed] = []
    seen: set[tuple[str, str]] = set()
    for seed in seeds:
        term = seed.term.strip()
        if not _is_glossary_seed_text(term):
            continue
        key = (term.casefold(), seed.category)
        if key in seen:
            continue
        seen.add(key)
        deduped.append(GalgameGlossarySeed(term, seed.category, seed.source, seed.context))
    return tuple(deduped)


def _metadata_int(metadata: Mapping[str, object], key: str) -> int | None:
    value = metadata.get(key)
    return value if isinstance(value, int) else None


def _looks_like_translatable_json_key(text: str) -> bool:
    source_text = text.strip()
    if not source_text:
        return False
    if any(ord(character) > 127 for character in source_text):
        return True
    if any(character.isspace() for character in source_text):
        return True
    if any(character in "、。！？…「」『』，。！？,.!?;:'\"()[]{}<>《》" for character in source_text):
        return True
    return len(source_text) >= 20 and _SIMPLE_IDENTIFIER_RE.fullmatch(source_text) is None


def _find_first_quoted_string(line: str) -> _QuotedString | None:
    for start_index, quote in enumerate(line):
        if quote not in {'"', "'"}:
            continue
        escaped = False
        for index in range(start_index + 1, len(line)):
            character = line[index]
            if escaped:
                escaped = False
                continue
            if character == "\\":
                escaped = True
                continue
            if character == quote:
                return _QuotedString(
                    start_quote=start_index,
                    end_quote=index,
                    quote=quote,
                    text=line[start_index + 1 : index],
                )
            if character in "\r\n":
                break
    return None


def _parse_renpy_translatable_line(line: str) -> _RenPyLine | None:
    stripped = line.strip()
    if not stripped or stripped.startswith(("#", "$")):
        return None
    quoted = _find_first_quoted_string(line)
    if quoted is None:
        return None

    prefix = line[: quoted.start_quote].strip()
    suffix = line[quoted.end_quote + 1 :].strip()
    if suffix and not suffix.startswith(("#", ":", "with ")):
        return None
    if prefix in {"old", "new"}:
        if prefix == "old" and not quoted.text.strip():
            return None
        return _RenPyLine(kind=prefix, text=quoted.text, speaker=None)
    if not quoted.text.strip():
        return None
    if prefix:
        speaker = _renpy_speaker_from_prefix(prefix)
        if speaker is None:
            return None
        return _RenPyLine(kind="dialogue", text=quoted.text, speaker=speaker)
    return _RenPyLine(kind="narration", text=quoted.text, speaker=None)


def _renpy_speaker_from_prefix(prefix: str) -> str | None:
    tokens = prefix.split()
    if not tokens or tokens[0] in _RENPY_COMMAND_PREFIXES:
        return None
    if not all(token.isidentifier() for token in tokens):
        return None
    return tokens[0]


def _find_following_renpy_new_line(lines: list[str], old_line_index: int) -> int | None:
    for line_index in range(old_line_index + 1, len(lines)):
        stripped = lines[line_index].strip()
        if not stripped or stripped.startswith("#"):
            continue
        parsed = _parse_renpy_translatable_line(lines[line_index])
        return line_index if parsed is not None and parsed.kind == "new" else None
    return None


def _replace_first_quoted_string(line: str, translation: str) -> str:
    quoted = _find_first_quoted_string(line)
    if quoted is None:
        return line
    escaped_translation = _escape_script_string(translation, quoted.quote)
    return f"{line[: quoted.start_quote + 1]}{escaped_translation}{line[quoted.end_quote:]}"


def _escape_script_string(text: str, quote: str) -> str:
    escaped = text.replace("\\", "\\\\").replace(quote, f"\\{quote}")
    return escaped.replace("\r\n", "\\n").replace("\r", "\\n").replace("\n", "\\n")


def _is_kag_dialogue_line(stripped_line: str) -> bool:
    if stripped_line.startswith(("@", "*", ";", "//")):
        return False
    return bool(_KAG_TAG_RE.sub("", stripped_line).strip())


def _replace_kag_line_text(line: str, translation: str) -> str:
    newline = ""
    if line.endswith("\r\n"):
        line = line[:-2]
        newline = "\r\n"
    elif line.endswith(("\n", "\r")):
        newline = line[-1]
        line = line[:-1]
    indent_length = len(line) - len(line.lstrip(" \t"))
    return f"{line[:indent_length]}{translation}{newline}"


def _rpg_maker_file_kind(relative_path: str) -> str | None:
    path = PurePosixPath(relative_path)
    if path.suffix.lower() != ".json" or not _rpg_maker_path_has_data_dir(relative_path):
        return None
    name = path.name
    if _RPG_MAKER_MAP_FILE_RE.fullmatch(name):
        return "map"
    normalized_name = name.lower()
    if normalized_name == "commonevents.json":
        return "common_events"
    if normalized_name == "troops.json":
        return "troops"
    return None


def _rpg_maker_path_has_data_dir(relative_path: str) -> bool:
    parts = [part.lower() for part in PurePosixPath(relative_path).parts]
    return len(parts) <= 1 or "data" in parts


def _rpg_maker_text_slots(kind: str, root: object) -> list[_RpgMakerTextSlot]:
    slots: list[_RpgMakerTextSlot] = []
    for context, json_path, command_list in _rpg_maker_command_lists(kind, root):
        pending_speaker: str | None = None
        for command_index, command in enumerate(command_list):
            if not isinstance(command, dict):
                pending_speaker = None
                continue
            command_mapping = cast("Mapping[str, object]", command)
            code = _rpg_maker_command_code(command_mapping)
            parameters = command_mapping.get("parameters")
            if code is None or not isinstance(parameters, list):
                pending_speaker = None
                continue
            parameter_values = cast("list[object]", parameters)
            command_path = f"{json_path}[{command_index}]"
            if code == 101:
                pending_speaker = _rpg_maker_parameter_text(parameter_values, 4)
                if pending_speaker is not None:
                    slots.append(
                        _RpgMakerTextSlot(
                            text=pending_speaker,
                            speaker=pending_speaker,
                            context=context,
                            metadata={"kind": "speaker", "code": code, "json_path": f"{command_path}.parameters[4]"},
                            parameters=parameter_values,
                            parameter_index=4,
                        )
                    )
                continue
            if code == 401:
                text = _rpg_maker_parameter_text(parameter_values, 0)
                if text is not None:
                    slots.append(
                        _RpgMakerTextSlot(
                            text=text,
                            speaker=pending_speaker,
                            context=context,
                            metadata={"kind": "dialogue", "code": code, "json_path": f"{command_path}.parameters[0]"},
                            parameters=parameter_values,
                            parameter_index=0,
                        )
                    )
                continue
            pending_speaker = None
            if code == 405:
                text = _rpg_maker_parameter_text(parameter_values, 0)
                if text is not None:
                    slots.append(
                        _RpgMakerTextSlot(
                            text=text,
                            speaker=None,
                            context=context,
                            metadata={"kind": "scroll_text", "code": code, "json_path": f"{command_path}.parameters[0]"},
                            parameters=parameter_values,
                            parameter_index=0,
                        )
                    )
                continue
            if code == 102:
                choices = parameter_values[0] if parameter_values else None
                if isinstance(choices, list):
                    for choice_index, choice in enumerate(choices):
                        if isinstance(choice, str) and choice.strip():
                            slots.append(
                                _RpgMakerTextSlot(
                                    text=choice,
                                    speaker=None,
                                    context=context,
                                    metadata={
                                        "kind": "choice",
                                        "code": code,
                                        "json_path": f"{command_path}.parameters[0][{choice_index}]",
                                    },
                                    parameters=parameter_values,
                                    parameter_index=0,
                                    choice_index=choice_index,
                                )
                            )
    return slots


def _rpg_maker_command_lists(kind: str, root: object) -> list[tuple[str, str, list[object]]]:
    if kind == "map":
        return _rpg_maker_map_command_lists(root)
    if kind == "common_events":
        return _rpg_maker_common_event_command_lists(root)
    if kind == "troops":
        return _rpg_maker_troop_command_lists(root)
    return []


def _rpg_maker_map_command_lists(root: object) -> list[tuple[str, str, list[object]]]:
    if not isinstance(root, dict):
        return []
    events = cast("Mapping[str, object]", root).get("events")
    if not isinstance(events, list):
        return []
    command_lists: list[tuple[str, str, list[object]]] = []
    for event_index, event in enumerate(events):
        if not isinstance(event, dict):
            continue
        event_mapping = cast("Mapping[str, object]", event)
        event_name = _optional_entry_str(event_mapping, "name") or f"event {event_index}"
        pages = event_mapping.get("pages")
        if not isinstance(pages, list):
            continue
        for page_index, page in enumerate(pages):
            if not isinstance(page, dict):
                continue
            commands = cast("Mapping[str, object]", page).get("list")
            if isinstance(commands, list):
                context = f"Map {event_name}, page {page_index + 1}"
                command_lists.append((context, f"events[{event_index}].pages[{page_index}].list", commands))
    return command_lists


def _rpg_maker_common_event_command_lists(root: object) -> list[tuple[str, str, list[object]]]:
    if not isinstance(root, list):
        return []
    command_lists: list[tuple[str, str, list[object]]] = []
    for event_index, event in enumerate(root):
        if not isinstance(event, dict):
            continue
        event_mapping = cast("Mapping[str, object]", event)
        commands = event_mapping.get("list")
        if isinstance(commands, list):
            event_name = _optional_entry_str(event_mapping, "name") or f"common event {event_index}"
            command_lists.append((f"Common event {event_name}", f"[{event_index}].list", commands))
    return command_lists


def _rpg_maker_troop_command_lists(root: object) -> list[tuple[str, str, list[object]]]:
    if not isinstance(root, list):
        return []
    command_lists: list[tuple[str, str, list[object]]] = []
    for troop_index, troop in enumerate(root):
        if not isinstance(troop, dict):
            continue
        troop_mapping = cast("Mapping[str, object]", troop)
        troop_name = _optional_entry_str(troop_mapping, "name") or f"troop {troop_index}"
        pages = troop_mapping.get("pages")
        if not isinstance(pages, list):
            continue
        for page_index, page in enumerate(pages):
            if not isinstance(page, dict):
                continue
            commands = cast("Mapping[str, object]", page).get("list")
            if isinstance(commands, list):
                context = f"Troop {troop_name}, page {page_index + 1}"
                command_lists.append((context, f"[{troop_index}].pages[{page_index}].list", commands))
    return command_lists


def _rpg_maker_command_code(command: Mapping[str, object]) -> int | None:
    code = command.get("code")
    return code if isinstance(code, int) else None


def _rpg_maker_parameter_text(parameters: list[object], index: int) -> str | None:
    if index >= len(parameters):
        return None
    value = parameters[index]
    return value if isinstance(value, str) and value.strip() else None


def _patch_rpg_maker_slot(slot: _RpgMakerTextSlot, translation: str) -> None:
    if slot.choice_index is None:
        slot.parameters[slot.parameter_index] = translation
        return
    choices = slot.parameters[slot.parameter_index]
    if isinstance(choices, list) and 0 <= slot.choice_index < len(choices):
        choices[slot.choice_index] = translation


def _normalize_spreadsheet_header(value: object) -> str:
    if not isinstance(value, str):
        return ""
    normalized = value.strip().lower()
    for token in (" ", "_", "-", "(", ")", "[", "]", "（", "）"):
        normalized = normalized.replace(token, "")
    return normalized


def _first_matching_header(header_map: Mapping[str, int], candidates: Iterable[str]) -> int | None:
    for candidate in candidates:
        column = header_map.get(candidate)
        if column is not None:
            return column
    return None


def _spreadsheet_source_text(value: object) -> str | None:
    if not isinstance(value, str):
        return None
    text = value.strip()
    return text or None


def _spreadsheet_optional_text(value: object) -> str | None:
    if value is None:
        return None
    if isinstance(value, str):
        text = value.strip()
        return text or None
    return str(value)
